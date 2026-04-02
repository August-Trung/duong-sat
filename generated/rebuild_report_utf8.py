from pathlib import Path
from datetime import datetime
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


WORKSPACE = Path(r"D:\Study\Projects\DuongSat")
DB_PATH = WORKSPACE / "data" / "railway.db"
OUT_PATH = WORKSPACE / "generated" / "bao-cao-chuc-nang-admin-user-utf8.xlsx"


def vi(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


def load_counts() -> dict[str, int]:
    conn = sqlite3.connect(DB_PATH)
    try:
        tables = [
            "crossings",
            "train_schedules",
            "incident_reports",
            "news_articles",
            "users",
            "crossing_images",
            "risk_snapshots",
            "audit_logs",
        ]
        return {table: conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0] for table in tables}
    finally:
        conn.close()


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[tuple]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    pass_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FDE9D9")
    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    body_font = Font(name="Arial")
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="C9D2D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws = wb.create_sheet(title)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
        cell.border = border

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = wrap
            cell.border = border

        status_text = " ".join(str(part) for part in row)
        row_fill = warn_fill if ("Ch\u01b0a h\u1ed7 tr\u1ee3" in status_text or "Kh\u00f4ng active" in status_text or "Not active" in status_text) else pass_fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = row_fill

    widths = [14, 20, 38, 18, 44, 92]
    for idx, width in enumerate(widths[: len(headers)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def main() -> None:
    counts = load_counts()
    wb = Workbook()

    header_font = Font(bold=True, name="Arial")
    body_font = Font(name="Arial")
    wrap = Alignment(vertical="top", wrap_text=True)

    ws = wb.active
    ws.title = "Tong_quan"
    ws["A1"] = vi("B\\u00e1o c\\u00e1o ch\\u1ee9c n\\u0103ng Admin/User \\u0111ang ch\\u1ea1y th\\u1eadt")
    ws["A2"] = vi(f"Ng\\u00e0y xu\\u1ea5t: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ws["A3"] = vi("Nguy\\u00ean t\\u1eafc ch\\u1ed1t danh s\\u00e1ch")
    principles = [
        vi("Ch\\u1ec9 ghi nh\\u1eadn ch\\u1ee9c n\\u0103ng c\\u00f3 UI/route ho\\u1eb7c API/backend th\\u1ef1c s\\u1ef1 t\\u1ed3n t\\u1ea1i trong code hi\\u1ec7n t\\u1ea1i."),
        vi("C\\u00e1c ch\\u1ee9c n\\u0103ng ghi d\\u1eef li\\u1ec7u quan tr\\u1ecdng \\u0111\\u00e3 \\u0111\\u01b0\\u1ee3c verify b\\u1eb1ng FastAPI TestClient tr\\u00ean DB copy trong workspace, kh\\u00f4ng \\u0111\\u1ee5ng DB th\\u1eadt."),
        vi("C\\u00e1c module cho ph\\u00e9p nh\\u1eadp tay d\\u1eef li\\u1ec7u v\\u1eabn \\u0111\\u01b0\\u1ee3c t\\u00ednh l\\u00e0 ch\\u1ee9c n\\u0103ng th\\u1eadt n\\u1ebfu c\\u00f3 lu\\u1ed3ng l\\u01b0u, s\\u1eeda, \\u0111\\u1ecdc v\\u00e0 hi\\u1ec3n th\\u1ecb th\\u1eadt."),
        vi("C\\u00e1c ph\\u1ea7n ch\\u1ec9 c\\u00f2n code n\\u1ec1n nh\\u01b0ng kh\\u00f4ng m\\u1edf route ho\\u1eb7c API 404 \\u0111\\u01b0\\u1ee3c \\u0111\\u01b0a sang sheet Khong_active, kh\\u00f4ng t\\u00ednh l\\u00e0 ch\\u1ee9c n\\u0103ng \\u0111ang ch\\u1ea1y."),
    ]
    for idx, text in enumerate(principles, start=4):
        ws[f"A{idx}"] = f"- {text}"

    ws["A10"] = vi("S\\u1ed1 li\\u1ec7u DB hi\\u1ec7n c\\u00f3")
    summary_rows = [
        ("Crossings", counts["crossings"]),
        ("Train schedules", counts["train_schedules"]),
        ("Incident reports", counts["incident_reports"]),
        ("News articles", counts["news_articles"]),
        ("Users", counts["users"]),
        ("Crossing images", counts["crossing_images"]),
        ("Risk snapshots", counts["risk_snapshots"]),
        ("Audit logs", counts["audit_logs"]),
    ]
    for idx, (label, value) in enumerate(summary_rows, start=11):
        ws[f"A{idx}"] = label
        ws[f"B{idx}"] = value

    ws["A21"] = vi("K\\u1ebft lu\\u1eadn nhanh")
    ws["A22"] = vi("Public/user hi\\u1ec7n c\\u00f3 \\u0111\\u1ee7 lu\\u1ed3ng tra c\\u1ee9u 2D, l\\u1ecdc, \\u0111\\u1ecbnh v\\u1ecb, h\\u1ed3 s\\u01a1 \\u0111i\\u1ec3m, danh m\\u1ee5c v\\u00e0 insights.")
    ws["A23"] = vi("Admin hi\\u1ec7n c\\u00f3 lu\\u1ed3ng v\\u1eadn h\\u00e0nh th\\u1eadt cho dashboard, crossings, \\u1ea3nh hi\\u1ec7n tr\\u01b0\\u1eddng, schedules, incidents, users, export CSV.")
    ws["A24"] = vi("Kh\\u00f4ng t\\u00ednh l\\u00e0 active: scene 3D public v\\u00e0 x\\u00f3a c\\u1ee9ng user.")

    for row in ws.iter_rows(min_row=1, max_row=24, min_col=1, max_col=2):
        for cell in row:
            cell.font = body_font
            cell.alignment = wrap
    for cell_ref in ["A1", "A3", "A10", "A21"]:
        ws[cell_ref].font = header_font
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18

    user_rows = [
        (
            "User",
            vi("\\u0110i\\u1ec1u h\\u01b0\\u1edbng v\\u00e0 tra c\\u1ee9u"),
            vi("Thanh \\u0111i\\u1ec1u h\\u01b0\\u1edbng public v\\u00e0 t\\u1ea3i d\\u1eef li\\u1ec7u t\\u1ed5ng quan"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            "UI public hoat dong; goi /api/summary, /api/layers, /api/crossings, /api/schedules, /api/incidents",
            vi("Trang public t\\u1ef1 load d\\u1eef li\\u1ec7u t\\u1ed5ng quan, s\\u1ed1 li\\u1ec7u metric v\\u00e0 3 tab B\\u1ea3n \\u0111\\u1ed3, Danh m\\u1ee5c, C\\u1ea3nh b\\u00e1o. D\\u1eef li\\u1ec7u l\\u1ea5y tr\\u1ef1c ti\\u1ebfp t\\u1eeb SQLite hi\\u1ec7n c\\u00f3."),
        ),
        (
            "User",
            vi("T\\u00ecm ki\\u1ebfm"),
            vi("T\\u00ecm nhanh \\u0111i\\u1ec3m giao c\\u1eaft v\\u00e0 g\\u1ee3i \\u00fd t\\u1ee9c th\\u00ec"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("UI t\\u00ecm theo t\\u00ean, m\\u00e3, \\u0111\\u1ecba ch\\u1ec9; g\\u1ee3i \\u00fd sinh t\\u1eeb dataset th\\u1eadt"),
            vi("\\u00d4 t\\u00ecm ki\\u1ebfm l\\u1ecdc tr\\u00ean t\\u1eadp crossings th\\u1eadt v\\u00e0 hi\\u1ec7n suggestion ngay khi g\\u00f5."),
        ),
        (
            "User",
            vi("B\\u1ed9 l\\u1ecdc"),
            vi("L\\u1ecdc theo m\\u1ee9c nguy hi\\u1ec3m, r\\u00e0o ch\\u1eafn, khu v\\u1ef1c, s\\u1eafp x\\u1ebfp"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("K\\u1ebft h\\u1ee3p filter API v\\u00e0 filter client"),
            vi("Ng\\u01b0\\u1eddi d\\u00f9ng c\\u00f3 th\\u1ec3 l\\u1ecdc risk level, barrier type, district, kho\\u1ea3ng c\\u00e1ch v\\u00e0 sort tr\\u00ean d\\u1eef li\\u1ec7u th\\u1eadt."),
        ),
        (
            "User",
            vi("V\\u1ecb tr\\u00ed"),
            vi("D\\u00f9ng v\\u1ecb tr\\u00ed hi\\u1ec7n t\\u1ea1i \\u0111\\u1ec3 t\\u00ednh kho\\u1ea3ng c\\u00e1ch v\\u00e0 g\\u1ee3i \\u00fd g\\u1ea7n t\\u00f4i"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("D\\u00f9ng navigator.geolocation c\\u1ee7a tr\\u00ecnh duy\\u1ec7t"),
            vi("Khi user c\\u1ea5p quy\\u1ec1n v\\u1ecb tr\\u00ed, h\\u1ec7 th\\u1ed1ng l\\u1ea5y GPS th\\u1eadt \\u0111\\u1ec3 t\\u00ednh nearby v\\u00e0 v\\u00f9ng quan t\\u00e2m."),
        ),
    ]

    admin_rows = [
        (
            "Admin",
            vi("X\\u00e1c th\\u1ef1c"),
            vi("\\u0110\\u0103ng nh\\u1eadp, \\u0111\\u0103ng xu\\u1ea5t v\\u00e0 guard khu v\\u1ef1c qu\\u1ea3n tr\\u1ecb"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("\\u0110\\u00e3 verify /api/auth/login v\\u00e0 /api/auth/me"),
            vi("Backend t\\u1ea1o token phi\\u00ean trong auth_sessions v\\u00e0 frontend guard route /admin."),
        ),
        (
            "Admin",
            "Dashboard",
            vi("T\\u1ed5ng quan v\\u1eadn h\\u00e0nh admin"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("GET /api/admin/overview tr\\u1ea3 d\\u1eef li\\u1ec7u th\\u1eadt"),
            vi("M\\u00e0n h\\u00ecnh t\\u1ed5ng quan load crossings, schedules, incidents, users, quality alerts, audit logs v\\u00e0 permission matrix."),
        ),
        (
            "Admin",
            vi("\\u0110i\\u1ec3m giao c\\u1eaft"),
            vi("T\\u1ea1o, s\\u1eeda, \\u1ea9n m\\u1ec1m v\\u00e0 kh\\u00f4i ph\\u1ee5c \\u0111i\\u1ec3m giao c\\u1eaft"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("\\u0110\\u00e3 test tr\\u00ean DB copy"),
            vi("Form admin ghi th\\u1eadt xu\\u1ed1ng b\\u1ea3ng crossings; x\\u00f3a l\\u00e0 soft delete qua deleted_at."),
        ),
        (
            "Admin",
            vi("\\u1ea2nh hi\\u1ec7n tr\\u01b0\\u1eddng"),
            vi("Upload, s\\u1eafp x\\u1ebfp, ch\\u1ecdn \\u1ea3nh \\u0111\\u1ea1i di\\u1ec7n, x\\u00f3a \\u1ea3nh"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("\\u0110\\u00e3 test /api/admin/crossings/{id}/images"),
            vi("\\u1ea2nh l\\u01b0u th\\u1eadt trong data/uploads, metadata l\\u01b0u trong crossing_images, c\\u00f3 audit log."),
        ),
        (
            "Admin",
            vi("Gi\\u1edd t\\u00e0u"),
            vi("T\\u1ea1o, s\\u1eeda, \\u1ea9n l\\u1ecbch t\\u00e0u"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("\\u0110\\u00e3 test schedule CRUD"),
            vi("CRUD l\\u1ecbch t\\u00e0u ho\\u1ea1t \\u0111\\u1ed9ng th\\u1eadt; d\\u1eef li\\u1ec7u c\\u00f3 th\\u1ec3 nh\\u1eadp tay."),
        ),
        (
            "Admin",
            vi("S\\u1ef1 c\\u1ed1"),
            vi("T\\u1ea1o, s\\u1eeda, \\u1ea9n s\\u1ef1 c\\u1ed1"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("\\u0110\\u00e3 test incident CRUD"),
            vi("CRUD s\\u1ef1 c\\u1ed1 ho\\u1ea1t \\u0111\\u1ed9ng th\\u1eadt v\\u00e0 l\\u01b0u trong incident_reports."),
        ),
        (
            "Admin",
            vi("Ng\\u01b0\\u1eddi d\\u00f9ng"),
            vi("T\\u1ea1o user, c\\u1eadp nh\\u1eadt role, tr\\u1ea1ng th\\u00e1i ho\\u1ea1t \\u0111\\u1ed9ng v\\u00e0 m\\u1eadt kh\\u1ea9u"),
            vi("\\u0110ang ch\\u1ea1y th\\u1eadt"),
            vi("\\u0110\\u00e3 test /api/admin/users"),
            vi("Backend hash m\\u1eadt kh\\u1ea9u th\\u1eadt b\\u1eb1ng PBKDF2; x\\u00f3a c\\u1ee9ng user ch\\u01b0a h\\u1ed7 tr\\u1ee3."),
        ),
        (
            "Admin",
            vi("Ng\\u01b0\\u1eddi d\\u00f9ng"),
            vi("X\\u00f3a c\\u1ee9ng user"),
            vi("Ch\\u01b0a h\\u1ed7 tr\\u1ee3"),
            vi("UI ch\\u1eb7n v\\u00e0 backend kh\\u00f4ng c\\u00f3 DELETE /api/admin/users/{id}"),
            vi("Lu\\u1ed3ng thay th\\u1ebf l\\u00e0 chuy\\u1ec3n sang ng\\u1eebng ho\\u1ea1t \\u0111\\u1ed9ng qua is_active."),
        ),
    ]

    inactive_rows = [
        (
            vi("C\\u00f4ng khai"),
            "Scene 3D public",
            vi("Kh\\u00f4ng active"),
            vi("Router hi\\u1ec7n redirect /scene-3d v\\u1ec1 b\\u1ea3n \\u0111\\u1ed3 th\\u01b0\\u1eddng v\\u00e0 backend /api/scene3d/manifest tr\\u1ea3 404."),
        ),
        (
            "Admin",
            vi("X\\u00f3a c\\u1ee9ng user"),
            vi("Kh\\u00f4ng active"),
            vi("Ph\\u1ea7n x\\u00f3a user kh\\u00f4ng \\u0111\\u01b0\\u1ee3c tri\\u1ec3n khai th\\u1eadt."),
        ),
    ]

    validation_rows = [
        ("Public summary", "/api/summary", 200, "Pass", vi("\\u0110\\u1ecdc s\\u1ed1 li\\u1ec7u t\\u1ed5ng quan c\\u00f4ng khai t\\u1eeb DB th\\u1eadt")),
        ("Auth login", "/api/auth/login", 200, "Pass", vi("\\u0110\\u0103ng nh\\u1eadp admin/admin123 th\\u00e0nh c\\u00f4ng")),
        ("Admin overview", "/api/admin/overview", 200, "Pass", vi("Dashboard admin l\\u1ea5y d\\u1eef li\\u1ec7u th\\u1eadt")),
        ("Crossing profile", "/api/admin/crossings/13/profile", 200, "Pass", vi("H\\u1ed3 s\\u01a1 360 ho\\u1ea1t \\u0111\\u1ed9ng")),
        ("Scene 3D manifest", "/api/scene3d/manifest", 404, "Not active", vi("Endpoint scene 3D hi\\u1ec7n kh\\u00f4ng mount")),
    ]

    add_sheet(
        wb,
        "User",
        [vi("Vai tr\\u00f2"), vi("Nh\\u00f3m"), vi("Ch\\u1ee9c n\\u0103ng"), vi("Tr\\u1ea1ng th\\u00e1i"), vi("B\\u1eb1ng ch\\u1ee9ng"), vi("M\\u00f4 t\\u1ea3 chi ti\\u1ebft")],
        user_rows,
    )
    add_sheet(
        wb,
        "Admin",
        [vi("Vai tr\\u00f2"), vi("Nh\\u00f3m"), vi("Ch\\u1ee9c n\\u0103ng"), vi("Tr\\u1ea1ng th\\u00e1i"), vi("B\\u1eb1ng ch\\u1ee9ng"), vi("M\\u00f4 t\\u1ea3 chi ti\\u1ebft")],
        admin_rows,
    )
    add_sheet(
        wb,
        "Khong_active",
        [vi("Ph\\u1ea1m vi"), vi("Ch\\u1ee9c n\\u0103ng"), vi("Tr\\u1ea1ng th\\u00e1i"), vi("Ghi ch\\u00fa")],
        inactive_rows,
    )
    add_sheet(
        wb,
        "Validation",
        [vi("H\\u1ea1ng m\\u1ee5c test"), "Endpoint", "HTTP", vi("K\\u1ebft qu\\u1ea3"), vi("Ghi ch\\u00fa")],
        validation_rows,
    )

    wb.save(OUT_PATH)
    print(OUT_PATH)


if __name__ == "__main__":
    main()
